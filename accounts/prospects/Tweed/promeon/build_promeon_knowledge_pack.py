from pathlib import Path
from openpyxl import load_workbook

WORKBOOK = Path("/Users/letiii/Documents/New project/outputs/promeon_4_0/PROMEON Version 4.0-05.09.2026.xlsx")
OUT = Path("/Users/letiii/Documents/New project/outputs/promeon_4_0/knowledge_pack")


def clean(value):
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def table_to_markdown(ws, max_rows=None):
    rows = list(ws.iter_rows(values_only=True))
    if max_rows:
        rows = rows[:max_rows]
    while rows and not any(clean(v) for v in rows[-1]):
        rows.pop()
    if not rows:
        return ""
    width = max(len(r) for r in rows)
    normalized = []
    for row in rows:
        normalized.append([clean(v).replace("|", "\\|") for v in list(row) + [""] * (width - len(row))])
    header = normalized[0]
    body = normalized[1:]
    lines = ["| " + " | ".join(header) + " |", "| " + " | ".join(["---"] * width) + " |"]
    for row in body:
        if any(row):
            lines.append("| " + " | ".join(row) + " |")
    return "\n".join(lines)


def write(path, text):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def build():
    OUT.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)

    prompt = clean(wb["System_Prompt_v4"]["A4"].value)
    write(
        OUT / "PROMEON_4_0_Constitution.md",
        "# PROMEON 4.0 Constitution\n\n"
        "Use this file as the core behavioral constitution for PROMEON. It defines identity, reasoning discipline, facilitation logic, uncertainty governance, ontology use, and deliverable standards.\n\n"
        + prompt
        + "\n",
    )

    sections = [
        ("Diagnostic_Workflow_v4", "Required Diagnostic Workflow"),
        ("Reasoning_Primitives_v4", "Reasoning Primitives"),
        ("Anti_Patterns_v4", "Anti-Pattern Suppression"),
        ("System_State_Model_v4", "System State Model"),
        ("Validation_Tests_v4", "Validation Test Suite"),
        ("Prompt_Templates_v4", "Prompt Templates"),
    ]
    runtime_parts = ["# PROMEON 4.0 Runtime Protocols\n"]
    for sheet, title in sections:
        runtime_parts.append(f"\n## {title}\n\n")
        runtime_parts.append(table_to_markdown(wb[sheet]))
        runtime_parts.append("\n")
    write(OUT / "PROMEON_4_0_Runtime_Protocols.md", "\n".join(runtime_parts))

    ontology_parts = ["# PROMEON Ontology Core\n"]
    for sheet, title, rows in [
        ("PROM_Terms", "PROM Terms", None),
        ("Relation_Types", "Relation Types", None),
        ("Entity_Types", "Entity Types", None),
        ("Actor_Ontology_v3", "Actor Ontology", None),
        ("Mechanisms_Normalized", "Mechanisms Normalized", None),
    ]:
        ontology_parts.append(f"\n## {title}\n\n")
        ontology_parts.append(table_to_markdown(wb[sheet], rows))
        ontology_parts.append("\n")
    write(OUT / "PROMEON_Ontology_Core.md", "\n".join(ontology_parts))

    graph_parts = ["# PROMEON Graph and Evidence Core\n"]
    for sheet, title, rows in [
        ("Edges_Production", "Edges Production", None),
        ("Failure_Patterns", "Failure Patterns", None),
        ("Diagnostics_Operational", "Operational Diagnostics", None),
        ("Evidence", "Evidence", None),
        ("Cases", "Cases", None),
        ("Contexts", "Contexts", None),
    ]:
        graph_parts.append(f"\n## {title}\n\n")
        graph_parts.append(table_to_markdown(wb[sheet], rows))
        graph_parts.append("\n")
    write(OUT / "PROMEON_Graph_Evidence_Core.md", "\n".join(graph_parts))

    usage = """# PROMEON 4.0 GPT Knowledge Upload Guide

Upload these files to the Custom GPT Knowledge section:

1. PROMEON_4_0_Constitution.md
2. PROMEON_4_0_Runtime_Protocols.md
3. PROMEON_Ontology_Core.md
4. PROMEON_Graph_Evidence_Core.md

Optional:
- PROMEON Version 4.0-05.09.2026.xlsx as a complete editable source archive.

Use the markdown files as the primary knowledge base because they are more text-forward and easier for GPT Knowledge retrieval to chunk. The workbook can be uploaded as a backup reference, but do not rely on it alone.
"""
    write(OUT / "UPLOAD_THESE_FILES.md", usage)

    for p in sorted(OUT.iterdir()):
        print(p)


if __name__ == "__main__":
    build()
